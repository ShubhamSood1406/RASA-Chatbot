import sqlite3
from openpyxl import load_workbook

conn = sqlite3.connect('events_list.db')
wb = load_workbook('Events_Details.xlsx')
ws = wb['events']

conn.execute("create table if not exists event_items (Cities text, Genre text, Events text, Date_Time text, Price int)")

for i in range(1,63):
    temp_str = "insert into event_items (Cities, Genre, Events, Date_Time, Price) values ('{0}', '{1}', '{2}', '{3}', '{4}')".format(ws.cell(i,1).value, ws.cell(i,2).value, ws.cell(i,3).value, ws.cell(i,4).value, ws.cell(i,5).value)
    conn.execute(temp_str)

conn.commit()

content = conn.execute("select * from event_items")
for i in content:
    print(i)
